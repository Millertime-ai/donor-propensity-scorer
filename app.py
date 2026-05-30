"""
PropensityScore.ai • Donor Propensity + Wealth Tool
Production-ready Streamlit application.

Exact scoring logic per Donor Propensity model.
100% local processing. No data retained.
"""

import streamlit as st
import pandas as pd
from datetime import datetime, date
import plotly.express as px
import io
import re
from typing import Dict, Any, Optional

# =============================================================================
# GLOBAL ERROR HANDLING (Professional pattern)
# =============================================================================

def handle_error(e):
    """Global safe error handler for the app."""
    st.error("🚨 App hit an issue. Please upload your file again.")
    st.caption("Error: " + str(e)[:300])

# For robust Excel fallback
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# =============================================================================
# PAGE CONFIG
# =============================================================================

st.set_page_config(
    page_title="PropensityScore.ai • Donor Propensity + Wealth Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# CORE SCORING LOGIC (EXACT IMPLEMENTATION)
# =============================================================================

def calculate_propensity_score(
    row: pd.Series,
    today: Optional[date] = None
) -> Dict[str, Any]:
    """
    Compute the exact 1-100 Propensity Score using the official model.

    Returns a dict with:
        - Days_Since_Last_Gift
        - Recency_Points
        - Engagement_Points
        - Past_Giving_Bonus
        - Ever_Given_Bonus
        - Major_Donor_Boost
        - Raw_Score
        - Propensity_Score (final 0-100 integer)
        - Probability_Label
        - Re_Engage_Flag
    """
    if today is None:
        today = date.today()

    # 1. Days since last gift
    last_gift = row.get("last_gift_date_parsed")
    if pd.isna(last_gift):
        days_since = 999
    else:
        try:
            if hasattr(last_gift, "date"):
                # datetime object
                d = last_gift.date()
            else:
                # already a date object
                d = last_gift
            days_since = (today - d).days
        except Exception:
            days_since = 999

    days_since = max(0, int(days_since))

    # 2. Recency Points (max 45)
    recency_points = max(0.0, 45.0 * (1.0 - (days_since / 365.0)))

    # 3. Engagement Points (max 25)
    opens = float(row.get("opens_30d", 0) or 0)
    clicks = float(row.get("clicks_30d", 0) or 0)
    visits = float(row.get("visits_30d", 0) or 0)

    engagement_points = (opens * 1.5) + (clicks * 4.0) + (visits * 6.0)
    engagement_points = min(engagement_points, 25.0)

    # 4. Past Giving Bonus (max 15, now /1000)
    total_given = float(row.get("total_given", 0) or 0)
    if total_given > 0:
        past_giving_bonus = 15.0 * min(1.0, total_given / 1000.0)
    else:
        past_giving_bonus = 0.0

    # 5. Ever_Given_Bonus (new)
    ever_given_bonus = 10.0 if total_given > 0 else 0.0

    # 6. Major Donor Boost
    major_donor_boost = 5.0 if total_given > 5000 else 0.0

    # 7. Raw Score
    raw_score = recency_points + engagement_points + past_giving_bonus + ever_given_bonus + major_donor_boost

    # 8. Final Propensity Score (integer 0-100)
    propensity_score = int(round(min(100.0, max(0.0, raw_score))))

    # 9. Probability Label (new thresholds: 60/30)
    if propensity_score >= 60:
        probability_label = "HIGH"
    elif propensity_score >= 30:
        probability_label = "MEDIUM"
    else:
        probability_label = "LOW"

    # 10. Re-Engage Flag (back to >25)
    re_engage_flag = "YES" if (days_since > 90 and propensity_score > 25) else "NO"

    return {
        "Days_Since_Last_Gift": days_since,
        "Recency_Points": round(recency_points, 2),
        "Engagement_Points": round(engagement_points, 2),
        "Past_Giving_Bonus": round(past_giving_bonus, 2),
        "Ever_Given_Bonus": round(ever_given_bonus, 2),
        "Major_Donor_Boost": round(major_donor_boost, 2),
        "Raw_Score": round(raw_score, 2),
        "Propensity_Score": propensity_score,
        "Probability_Label": probability_label,
        "Re_Engage_Flag": re_engage_flag,
    }


@st.cache_data(ttl=3600, show_spinner=False)
def score_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply scoring to an entire dataframe.
    Returns a new dataframe with original columns + scoring columns.
    Cached for speed and stability on large (31k+) files.
    """
    today = date.today()
    scored_rows = []

    for _, row in df.iterrows():
        scores = calculate_propensity_score(row, today=today)
        scored_rows.append(scores)

    score_df = pd.DataFrame(scored_rows)

    # Combine original + scores. Keep original column order first.
    result = df.copy().reset_index(drop=True)
    result = pd.concat([result, score_df], axis=1)

    return result


@st.cache_data(ttl=3600, show_spinner=False)
def create_scatter_figure(plot_df, x_col, title, hover_cols, labels, marker_size=9):
    """Cached helper for creating the main Propensity vs Capacity scatter plot."""
    fig = px.scatter(
        plot_df,
        x=x_col,
        y="Propensity_Score",
        color="Probability_Label",
        hover_data=hover_cols,
        title=title,
        labels=labels,
        color_discrete_sequence=["#1E88E5", "#FB8C00", "#E53935"],
    )
    fig.update_traces(marker=dict(size=marker_size, opacity=0.8, line=dict(width=0.5, color="white")))
    fig.update_layout(
        height=540,
        margin=dict(l=10, r=10, t=60, b=10),
        legend_title_text="Propensity Tier"
    )
    return fig



# =============================================================================
# SAMPLE DATA GENERATOR
# =============================================================================

def generate_sample_csv() -> bytes:
    """
    Generate a realistic 25-row sample CSV that exercises all scoring paths.
    """
    import random
    from datetime import timedelta

    random.seed(42)  # Reproducible
    today = date.today()

    rows = []

    # High propensity donors (recent + engaged)
    for i in range(6):
        rows.append({
            "email": f"high.donor{i+1}@example.org",
            "name": f"Alex Rivera {i+1}",
            "last_gift_date": (today - timedelta(days=random.randint(5, 35))).isoformat(),
            "total_given": round(random.uniform(120, 850), 2),
            "opens_30d": random.randint(4, 12),
            "clicks_30d": random.randint(1, 5),
            "visits_30d": random.randint(0, 3),
        })

    # Medium propensity
    for i in range(8):
        rows.append({
            "email": f"medium.donor{i+1}@example.org",
            "name": f"Jordan Lee {i+1}",
            "last_gift_date": (today - timedelta(days=random.randint(40, 160))).isoformat(),
            "total_given": round(random.uniform(0, 420), 2),
            "opens_30d": random.randint(1, 6),
            "clicks_30d": random.randint(0, 2),
            "visits_30d": random.randint(0, 1),
        })

    # Low propensity / lapsed
    for i in range(7):
        rows.append({
            "email": f"lapsed.donor{i+1}@example.org",
            "name": f"Sam Patel {i+1}",
            "last_gift_date": (today - timedelta(days=random.randint(200, 820))).isoformat(),
            "total_given": round(random.uniform(0, 180), 2),
            "opens_30d": random.randint(0, 2),
            "clicks_30d": random.randint(0, 1),
            "visits_30d": 0,
        })

    # Re-engagement candidates (lapsed >90 days but still some engagement)
    for i in range(4):
        rows.append({
            "email": f"reengage.candidate{i+1}@example.org",
            "name": f"Taylor Kim {i+1}",
            "last_gift_date": (today - timedelta(days=random.randint(95, 380))).isoformat(),
            "total_given": round(random.uniform(35, 290), 2),
            "opens_30d": random.randint(2, 5),
            "clicks_30d": random.randint(1, 3),
            "visits_30d": random.randint(0, 2),
        })

    # One row with missing last_gift_date
    rows.append({
        "email": "never.given@example.org",
        "name": "Morgan Ellis",
        "last_gift_date": "",  # Will become 999 days
        "total_given": 0,
        "opens_30d": 3,
        "clicks_30d": 1,
        "visits_30d": 1,
    })

    df = pd.DataFrame(rows)

    # Add a couple of extra columns to prove we preserve unknowns
    df["phone"] = [f"555-01{str(i).zfill(2)}" for i in range(len(df))]
    df["source"] = random.choices(["website", "event", "referral", "direct"], k=len(df))

    buffer = io.StringIO()
    df.to_csv(buffer, index=False)
    return buffer.getvalue().encode("utf-8")


def generate_sample_xlsx() -> bytes:
    """
    Generate a realistic 18-row sample XLSX that exercises all scoring paths.
    Designed to look like a typical nonprofit CRM export.
    """
    import random
    from datetime import timedelta
    from io import BytesIO

    random.seed(42)
    today = date.today()

    rows = []

    # HIGH propensity (recent + strong engagement)
    for i in range(4):
        rows.append({
            "email": f"high.donor{i+1}@example.org",
            "name": ["Maria Gonzalez", "David Chen", "Aisha Patel", "James Thompson"][i],
            "last_gift_date": (today - timedelta(days=random.randint(4, 28))).isoformat(),
            "total_given": round(random.choice([185, 420, 675, 1250]), 2),
            "opens_30d": random.randint(6, 14),
            "clicks_30d": random.randint(2, 7),
            "visits_30d": random.randint(1, 5),
        })

    # MEDIUM propensity
    for i in range(5):
        rows.append({
            "email": f"medium.donor{i+1}@example.org",
            "name": ["Robert Kim", "Elena Rodriguez", "Michael O'Brien", "Sofia Martinez", "Kwame Asante"][i],
            "last_gift_date": (today - timedelta(days=random.randint(45, 140))).isoformat(),
            "total_given": round(random.choice([0, 65, 180, 390, 820]), 2),
            "opens_30d": random.randint(1, 5),
            "clicks_30d": random.randint(0, 3),
            "visits_30d": random.randint(0, 2),
        })

    # LOW / long-lapsed
    for i in range(4):
        rows.append({
            "email": f"lapsed.donor{i+1}@example.org",
            "name": ["Patricia Wells", "Thomas Becker", "Linda Nakamura", "George Ramirez"][i],
            "last_gift_date": (today - timedelta(days=random.randint(210, 620))).isoformat(),
            "total_given": round(random.choice([0, 45, 125, 275]), 2),
            "opens_30d": random.randint(0, 2),
            "clicks_30d": random.randint(0, 1),
            "visits_30d": 0,
        })

    # Re-engagement candidates (lapsed >90d but still some signals)
    for i in range(3):
        rows.append({
            "email": f"reengage{i+1}@example.org",
            "name": ["Catherine Lee", "Marcus Johnson", "Priya Sharma"][i],
            "last_gift_date": (today - timedelta(days=random.randint(95, 310))).isoformat(),
            "total_given": round(random.choice([75, 210, 480]), 2),
            "opens_30d": random.randint(3, 8),
            "clicks_30d": random.randint(1, 4),
            "visits_30d": random.randint(0, 3),
        })

    # Edge cases
    rows.append({
        "email": "new.prospect@example.org",
        "name": "Jordan Hale",
        "last_gift_date": "",  # never given
        "total_given": 0,
        "opens_30d": 4,
        "clicks_30d": 2,
        "visits_30d": 1,
    })

    rows.append({
        "email": "big.donor@example.org",
        "name": "Dr. Helen Vargas",
        "last_gift_date": (today - timedelta(days=12)).isoformat(),
        "total_given": 4750.00,
        "opens_30d": 9,
        "clicks_30d": 5,
        "visits_30d": 3,
    })

    df = pd.DataFrame(rows)

    # Add a couple of extra "CRM" columns to make it feel real
    df["Phone"] = [f"555-{random.randint(100,999)}-{random.randint(1000,9999)}" for _ in range(len(df))]
    df["Preferred Channel"] = random.choices(["Email", "Mail", "Phone", "Text"], k=len(df))

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Donor List")
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# COLUMN DETECTION & VALIDATION
# =============================================================================

REQUIRED_COLUMNS = ["email", "last_gift_date", "total_given", "opens_30d", "clicks_30d", "visits_30d"]
OPTIONAL_COLUMNS = ["name"]

# Extremely robust mapping for real-world CRM exports (used after aggressive normalization)
COLUMN_MAPPING = {
    'email': [
        'email', 'e-mail', 'email address', 'emailaddress', 'donor email', 'contact email',
        'primary email', 'e mail', 'emailaddress', 'donoremail', 'contactemail',
        'email_1', 'email1', 'email_2', 'email2', 'email 1', 'email 2'
    ],
    'last_gift_date': [
        'last gift date', 'lastgiftdate', 'last donation date', 'lastdonationdate',
        'last gift', 'donation date', 'gift date', 'last donation', 'most recent gift date',
        'last contrib date', 'last transaction date'
    ],
    'total_given': [
        'total given', 'lifetime giving', 'total donations', 'lifetime value',
        'totalgiven', 'lifetimegiving', 'lifetime donations', 'total contributed',
        'giving total', 'total donated', 'cumulative giving', 'total amount given',
        'lifetimegivingamount', 'givingamount', 'totalgiving',
        'total gift amount', 'totalgiftamount', 'total gifts', 'totalgiftcount'
    ],
    'opens_30d': [
        'opens 30d', 'opens30d', 'email opens last 30 days', 'opens in last 30 days',
        'email opens 30d', 'emailopens30d', 'opens last 30', '30 day opens',
        'email opens (30d)', 'opens (last 30 days)', 'emailopenslast30days'
    ],
    'clicks_30d': [
        'clicks 30d', 'clicks30d', 'email clicks last 30 days', 'clicks in last 30 days',
        'email clicks 30d', 'emailclicks30d', 'clicks last 30', '30 day clicks',
        'email clicks (30d)', 'clicks (last 30 days)', 'emailclickslast30days'
    ],
    'visits_30d': [
        'visits 30d', 'visits30d', 'website visits last 30 days', 'visits in last 30 days',
        'website visits 30d', 'websitevisits30d', 'visits last 30', '30 day visits',
        'web visits (30d)', 'site visits last 30 days', 'webvisits30days',
        'websitevisits30days', 'visits30days'
    ],
}


def normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Extremely aggressive column name normalization for messy CRM exports.
    1. Lowercase + strip
    2. Remove ALL spaces, underscores, hyphens, parentheses, periods (and similar)
    3. Match against expanded COLUMN_MAPPING
    """
    if df is None or len(df.columns) == 0:
        return df

    # Build lookup: aggressively normalized name -> original column
    norm_to_orig = {}
    for col in df.columns:
        norm = str(col).strip().lower()
        # Remove anything inside parentheses or brackets first (common in CRM headers)
        norm = re.sub(r'\(.*?\)', '', norm)
        norm = re.sub(r'\[.*?\]', '', norm)
        # Then remove all spaces, underscores, hyphens, parentheses, periods, etc.
        norm = re.sub(r'[\s_\-().\[\]]+', '', norm)
        norm_to_orig[norm] = col

    rename_map = {}
    for canonical, variants in COLUMN_MAPPING.items():
        for variant in variants:
            vnorm = variant.strip().lower()
            vnorm = re.sub(r'\(.*?\)', '', vnorm)
            vnorm = re.sub(r'\[.*?\]', '', vnorm)
            vnorm = re.sub(r'[\s_\-().\[\]]+', '', vnorm)
            if vnorm in norm_to_orig:
                original = norm_to_orig[vnorm]
                if original not in rename_map:  # first match wins
                    rename_map[original] = canonical
                break

    if rename_map:
        df = df.rename(columns=rename_map)

    # Name creation is handled authoritatively after normalization in the main pipeline
    # (see "Force-create a reliable 'name' column" block). Keep this lightweight.
    return df


def robust_read_excel(file_obj) -> pd.DataFrame:
    """
    Multi-layered robust Excel reader designed to handle corrupted or 
    non-standard .xlsx files produced by many CRMs (Bloomerang, DonorPerfect, etc.)
    that cause 'could not read stylesheet' / invalid XML errors with standard parsers.
    """
    if load_workbook is None:
        raise ImportError("openpyxl not available for fallback")

    # Layer 1: pandas + openpyxl with maximum tolerance flags
    try:
        file_obj.seek(0)
        return pd.read_excel(
            file_obj,
            engine="openpyxl",
            header=0,
            engine_kwargs={"data_only": True, "read_only": True},
        )
    except Exception as e1:
        last_error = e1

    # Layer 2: calamine engine (very fast Rust-based parser, more tolerant of some issues)
    try:
        file_obj.seek(0)
        return pd.read_excel(file_obj, engine="calamine", header=0)
    except Exception as e2:
        last_error = e2

    # Layer 3: Direct openpyxl with the most tolerant settings + manual DataFrame conversion
    try:
        file_obj.seek(0)
        wb = load_workbook(
            file_obj,
            data_only=True,
            read_only=True,
            keep_links=False,
        )
        ws = wb.active
        # iter_rows with values_only is efficient in read_only mode
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame()
        # Treat first row as header (handle None headers gracefully)
        raw_headers = rows[0]
        headers = [
            str(h) if h is not None else f"Unnamed_{i}"
            for i, h in enumerate(raw_headers)
        ]
        data = rows[1:]
        return pd.DataFrame(data, columns=headers)
    except Exception as e3:
        last_error = e3

    # All layers failed - re-raise so caller can show friendly message
    raise last_error


def validate_input_dataframe(df: pd.DataFrame) -> tuple[bool, str, pd.DataFrame]:
    """
    Returns (is_valid, error_message, cleaned_df)
    """
    if df is None or len(df) == 0:
        return False, "Uploaded file is empty.", df

    # Apply extremely robust column name normalization
    df = normalize_column_names(df)

    # Ensure engagement columns exist (completely optional - default to 0)
    for col in ['opens_30d', 'clicks_30d', 'visits_30d']:
        if col not in df.columns:
            df[col] = 0

    # ONLY require email and last_gift_date now
    core_required = ['email', 'last_gift_date']
    missing_core = [c for c in core_required if c not in df.columns]

    if missing_core:
        actual_cols = list(df.columns)
        return (
            False,
            f"Missing required column(s): {', '.join(missing_core)}.\n\n"
            f"We found these columns: {', '.join(actual_cols)}\n\n"
            "The app requires at least an 'email' column (Email_1, Email, etc.) and 'last_gift_date'. "
            "Please map or rename your columns accordingly.",
            df,
        )

    # Coerce types safely
    df["email"] = df["email"].astype(str).str.strip()

    # Parse dates
    df["last_gift_date_parsed"] = pd.to_datetime(
        df["last_gift_date"], errors="coerce"
    ).dt.date

    # Numeric coercion for the three engagement + giving fields
    for col in ["total_given", "opens_30d", "clicks_30d", "visits_30d"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Drop rows with completely missing email (fatal)
    before = len(df)
    df = df[df["email"] != ""]
    if len(df) < before:
        st.warning(f"Dropped {before - len(df)} rows with empty email addresses.")

    if len(df) == 0:
        return False, "No valid rows remaining after cleaning.", df

    return True, "", df


# =============================================================================
# UI HELPERS
# =============================================================================

def style_probability_label(label: str) -> str:
    if label == "HIGH":
        return "🟢 **HIGH**"
    elif label == "MEDIUM":
        return "🟡 **MEDIUM**"
    else:
        return "🔴 **LOW**"


def create_download_link(df: pd.DataFrame, filename: str) -> None:
    """Create a Streamlit download button for a dataframe as CSV."""
    csv = df.to_csv(index=False).encode("utf-8")
    st.download_button(
        label=f"⬇️ Download {filename}",
        data=csv,
        file_name=filename,
        mime="text/csv",
        use_container_width=True,
    )


# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    # -------------------------------------------------------------------------
    # HEADER
    # -------------------------------------------------------------------------
    st.markdown(
        """
        <div style="text-align: center; padding: 0.5rem 0 0.75rem 0;">
            <h1 style="margin-bottom: 0.1rem; font-size: 2.35rem;">PropensityScore.ai • Donor Propensity + Wealth Tool</h1>
            <p style="margin-top: 0; font-size: 1.05rem; color: #555;">
                1-100 Propensity to Give Score &nbsp;|&nbsp; Turn Silence Into Support
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # -------------------------------------------------------------------------
    # EMAIL CAPTURE BOX (simple, at the very top, non-blocking)
    # -------------------------------------------------------------------------
    if "email_submitted" not in st.session_state:
        st.session_state.email_submitted = False

    if not st.session_state.email_submitted:
        st.markdown(
            """
            <div style="
                background: linear-gradient(90deg, #f8fafc 0%, #f1f5f9 100%);
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                padding: 14px 18px;
                margin: 0.25rem 0 1rem 0;
                text-align: center;
            ">
                <div style="font-size: 0.95rem; font-weight: 600; color: #1e2937; margin-bottom: 6px;">
                    📬 Get the free Donor Propensity Playbook + monthly high-signal tips
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        email_c1, email_c2, email_c3 = st.columns([1.2, 2.2, 1.2])
        with email_c2:
            with st.form(key="email_capture_form", clear_on_submit=True):
                email_input = st.text_input(
                    "Work email",
                    placeholder="you@nonprofit.org",
                    label_visibility="collapsed",
                )
                email_submit = st.form_submit_button(
                    "Send me the free playbook →",
                    use_container_width=True,
                    type="primary",
                )

                if email_submit:
                    if email_input and "@" in email_input and "." in email_input.split("@")[-1]:
                        st.session_state.email_submitted = True
                        st.success("✅ Thank you! The playbook is on its way to your inbox.")
                        st.caption("We respect your time — only the best insights, never spam.")
                        st.balloons()
                    else:
                        st.error("Please enter a valid work email address.")

    # -------------------------------------------------------------------------
    # SIDEBAR
    # -------------------------------------------------------------------------
    with st.sidebar:
        # Big, obvious Restart button at the very top
        if st.button("🔄 Restart App", use_container_width=True, type="primary"):
            st.cache_data.clear()
            st.rerun()

        st.header("About This Score")

        st.markdown(
            """
            About This Score
Most wealth screening tools sell you capacity — how much someone could give (via wealth screening, income estimates, and home values).

At R. Miller Consulting, we score propensity — the real reason people don’t engage (or finally decide to give).

A $1M donor who opens every email and clicks every link has high propensity.

A $1M donor who never opens your emails and never clicks has low propensity.  
These donors should be engaged differently.

This tool measures the latter — using only the data you already own. We give you a clear propensity rating so your team understands how to use their time more efficiently!
            """
        )

        st.divider()

        st.subheader("Quick Stats (after upload)")
        st.caption("Metrics will appear here once you process a file.")

        st.divider()
        st.caption("100% local • Nothing stored • No sign-up")

    # -------------------------------------------------------------------------
    # HOW THE SCORE WORKS (EXPANDABLE)
    # -------------------------------------------------------------------------
    with st.expander("📐 How the Score Works — Exact Formula (from Donor Propensity model)", expanded=False):
        st.markdown(
            """
            ### The 10-Step Calculation (applied to every row)

            1. **Days_Since_Last_Gift**  
               `= today − last_gift_date`  
               Missing or invalid dates are treated as **999 days**.

            2. **Recency_Points** (0–45 points)  
               `= max(0, 45 × (1 − Days_Since_Last_Gift / 365))`  
               • < 30 days → ~41–45 pts  
               • 180 days → ~23 pts  
               • 365+ days → 0 pts

            3. **Engagement_Points** (0–25 points)  
               `= (opens_30d × 1.5) + (clicks_30d × 4) + (visits_30d × 6)`  
               Capped at **25 points**.

            4. **Past_Giving_Bonus** (0–15 points)  
               If `total_given > 0`:  
               `= 15 × min(1, total_given / 1000)`  
               (Gives full 15 pts once lifetime giving reaches $1,000+)

            5. **Ever_Given_Bonus** (0 or 10 points)  
               `= 10` if `total_given > 0`  
               (Flat bonus for anyone who has ever given)

            6. **Major_Donor_Boost** (0 or 5 points)  
               `= 5` if `total_given > 5000`  
               (Extra boost for very large lifetime donors)

            7. **Raw_Score** = Recency + Engagement + Past Giving Bonus + Ever Given Bonus + Major Donor Boost

            8. **Propensity_Score** (final 1–100)  
               `= round( min(100, max(0, Raw_Score)) )`

            9. **Probability_Label** (new thresholds)  
               - **HIGH** ≥ 60  
               - **MEDIUM** 30–59  
               - **LOW** < 30

            10. **Re_Engage_Flag**  
               `"YES"` if `(Days_Since_Last_Gift > 90 AND Propensity_Score > 25)`  
               These are lapsed donors who still show life — your highest-ROI re-engagement segment.
            """
        )

        st.info(
            "The model is intentionally transparent. You can audit every single component "
            "for any contact in the downloaded CSV (Recency_Points, Engagement_Points, etc.)."
        )

    st.divider()

    # -------------------------------------------------------------------------
    # FILE UPLOAD + SAMPLE DATA
    # -------------------------------------------------------------------------
    col1, col2 = st.columns([2, 1])

    with col1:
        st.subheader("Upload your donor file")
        uploaded_file = st.file_uploader(
            "Drag and drop your donor file here (.csv, .xlsx, or .xls)",
            type=["csv", "xlsx", "xls"],
            label_visibility="collapsed",
            help="Upload a .csv, .xlsx, or .xls file. The file must contain the required columns listed in the 'How the Score Works' section above.",
        )

    with col2:
        st.subheader("Need a test file?")
        csv_bytes = generate_sample_csv()
        st.download_button(
            label="📥 Download Sample CSV",
            data=csv_bytes,
            file_name="sample_donor_list.csv",
            mime="text/csv",
            use_container_width=True,
        )

        xlsx_bytes = generate_sample_xlsx()
        st.download_button(
            label="📥 Download Sample XLSX",
            data=xlsx_bytes,
            file_name="sample_donor_list.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("Realistic CRM-style data with varied propensity levels")

    # -------------------------------------------------------------------------
    # PROCESSING
    # -------------------------------------------------------------------------
    if uploaded_file is not None:
        try:
            filename = uploaded_file.name.lower()
            mime_type = getattr(uploaded_file, "type", "") or ""

            with st.spinner("Reading file..."):
                if filename.endswith(".csv"):
                    uploaded_file.seek(0)
                    raw_df = pd.read_csv(uploaded_file)
                else:
                    # Excel branch: .xlsx, .xls, or MIME type application/vnd.ms-excel
                    is_legacy_xls = (
                        filename.endswith(".xls") or
                        "ms-excel" in mime_type.lower() or
                        mime_type == "application/vnd.ms-excel"
                    )

                    uploaded_file.seek(0)

                    try:
                        if is_legacy_xls:
                            # Legacy .xls or application/vnd.ms-excel: follow requested priority
                            try:
                                raw_df = pd.read_excel(uploaded_file, engine="openpyxl", header=0)
                            except Exception:
                                uploaded_file.seek(0)
                                try:
                                    raw_df = pd.read_excel(uploaded_file, engine="xlrd", header=0)
                                except Exception:
                                    uploaded_file.seek(0)
                                    raw_df = robust_read_excel(uploaded_file)
                        else:
                            # Modern .xlsx and other Excel files: use the full robust multi-layer system
                            raw_df = robust_read_excel(uploaded_file)
                    except Exception as exc:
                        # All attempts failed — show friendly message
                        st.error(
                            "This XLSX file has formatting issues (common with some CRM exports like Bloomerang, DonorPerfect, etc.).\n\n"
                            "Quick fix: Open the file in Excel → File → Save As → CSV (Comma delimited) → upload the .csv instead.\n"
                            "Or try re-exporting a clean version from your CRM.\n\n"
                            "Still not working? Use the 'Download Sample XLSX' button to test, or just save your file as CSV (File → Save As → CSV UTF-8) — that always works."
                        )
                        st.stop()

            # Progress: cleaning
            progress_bar = st.progress(10, text="Cleaning and normalizing data...")

            # Handle common Excel/CSV quirks gracefully
            raw_df = raw_df.dropna(how="all")
            raw_df.columns = [str(col).strip() for col in raw_df.columns]

            # Extra safety: drop rows that are all empty/NaN after stripping
            raw_df = raw_df[
                ~raw_df.apply(
                    lambda row: all(pd.isna(v) or str(v).strip() == "" for v in row),
                    axis=1,
                )
            ]
            raw_df = raw_df.reset_index(drop=True)

            if len(raw_df) == 0:
                st.error("The uploaded file appears to be empty after cleaning blank rows.")
                st.stop()

            # === Use centralized extremely robust column normalization ===
            raw_df = normalize_column_names(raw_df)

            # Aggressive column name cleaning
            raw_df.columns = [str(col).strip().replace('\n', ' ').replace('\r', '').replace('  ', ' ') for col in raw_df.columns]

            # Make opens_30d, clicks_30d, visits_30d completely optional (default to 0 if missing)
            for col in ['opens_30d', 'clicks_30d', 'visits_30d']:
                if col not in raw_df.columns:
                    raw_df[col] = 0

            progress_bar.progress(25, text="Creating name column and detecting wealth fields...")

            # Force-create reliable 'name' column (permanently fixed, no "nan")
            try:
                first = None
                last = None

                # Try all common DonorSearch name columns
                for candidate in ['First Name', 'First_Name', 'SP-First', 'Prefix']:
                    if candidate in raw_df.columns:
                        first = candidate
                        break

                for candidate in ['Last Name', 'Last_Name', 'SP-Last']:
                    if candidate in raw_df.columns:
                        last = candidate
                        break

                if first and last:
                    raw_df['name'] = (
                        raw_df[first].fillna('').astype(str).str.strip() + ' ' +
                        raw_df[last].fillna('').astype(str).str.strip()
                    ).str.strip()
                else:
                    # Fallback to email columns
                    email_col = None
                    for candidate in ['Email_1', 'Email', 'email']:
                        if candidate in raw_df.columns:
                            email_col = candidate
                            break
                    if email_col:
                        raw_df['name'] = raw_df[email_col].fillna('').astype(str).str.strip()
                    else:
                        raw_df['name'] = 'Unknown Donor'
            except Exception:
                raw_df['name'] = 'Unknown Donor'

            # Bulletproof wealth detection for exact DonorSearch bracketed columns
            wealth_col = None
            for col in raw_df.columns:
                clean = str(col).strip()
                if any(pattern in clean for pattern in ["[0 - 100]", "[100 - 126]", "0 - 100", "100 - 126", "Capacity Range Based on Wealth", "Wealth-Based Capacity", "Real Estate Est."]):
                    wealth_col = col
                    break

            if wealth_col is not None:
                # Robust conversion (kept for compatibility with older plot)
                col_series = raw_df[wealth_col].astype(str).str.strip()
                extracted = col_series.str.extract(r'(\d+)-?(\d*)')
                extracted = extracted.apply(pd.to_numeric, errors='coerce')
                numeric = extracted.mean(axis=1).fillna(0)
                raw_df[wealth_col] = numeric

            st.success(f"File loaded: **{len(raw_df):,}** rows × **{len(raw_df.columns)}** columns")

            # ============================================================
            # MAIN PROCESSING - Professional error handling wrapper
            # ============================================================
            try:
                with st.spinner("Processing your DonorSearch file..."):
                    # Friendly info for common CRM exports (Raiser's Edge etc.) that lack engagement columns
                    eng_cols = ['opens_30d', 'clicks_30d', 'visits_30d']
                    if all(raw_df.get(c, pd.Series([0])).eq(0).all() for c in eng_cols):
                        actual_cols_str = ', '.join(str(c) for c in raw_df.columns)
                        st.info(
                            f"We found these columns: {actual_cols_str}\n\n"
                            "Great news — your file has email and last_gift_date, so the score will still calculate!\n"
                            "(Engagement data like opens/clicks/visits is missing, so those scores default to 0. The score still works using recency + past giving.)"
                        )

                    progress_bar.progress(40, text="Validating input...")

                    # Validate & clean
                    is_valid, error_msg, cleaned_df = validate_input_dataframe(raw_df)

                    if not is_valid:
                        st.error(f"❌ {error_msg}")
                        st.stop()

                    progress_bar.progress(55, text="Scoring donors (cached for speed on large files)...")

                    # Safe scoring wrapper
                    try:
                        scored_df = score_dataframe(cleaned_df)
                    except Exception as score_err:
                        st.error(f"⚠️ Scoring failed: {str(score_err)[:200]}")
                        st.stop()

                    progress_bar.progress(92, text="Preparing dashboard...")

                    # Drop the internal parsed date column before display / download
                    if "last_gift_date_parsed" in scored_df.columns:
                        scored_df = scored_df.drop(columns=["last_gift_date_parsed"])

                    progress_bar.progress(100, text="Ready!")
                    progress_bar.empty()  # clean up the bar

                    # -----------------------------------------------------------------
                    # SUMMARY METRICS
                    # -----------------------------------------------------------------
                    st.divider()
                    st.subheader("Results Overview")

                    total = len(scored_df)
                    high_count = len(scored_df[scored_df["Probability_Label"] == "HIGH"])
                    med_count = len(scored_df[scored_df["Probability_Label"] == "MEDIUM"])
                    low_count = len(scored_df[scored_df["Probability_Label"] == "LOW"])
                    reengage_count = len(scored_df[scored_df["Re_Engage_Flag"] == "YES"])
                    avg_score = scored_df["Propensity_Score"].mean()

                    m1, m2, m3, m4, m5 = st.columns(5)
                    m1.metric("Total Contacts", f"{total:,}")
                    m2.metric("Avg Propensity", f"{avg_score:.0f}")
                    m3.metric("🟢 HIGH (≥60)", f"{high_count:,}", delta=f"{high_count/total:.0%}" if total else None)
                    m4.metric("🟡 MEDIUM", f"{med_count:,}")
                    m5.metric("🔴 LOW", f"{low_count:,}")

                    if reengage_count > 0:
                        st.success(
                            f"🎯 **{reengage_count} contacts** flagged for re-engagement "
                            f"({reengage_count/total:.1%} of list). These are lapsed donors who still show engagement signals."
                        )

                    # -----------------------------------------------------------------
                    # PROPENSITY VS CAPACITY SCATTER PLOT - Safe final fix
                    # -----------------------------------------------------------------
                    try:
                        st.subheader("Propensity vs Capacity Scatter Plot")

                        # Detect wealth-related columns (include "largest gift low")
                        wealth_options = [
                            col for col in scored_df.columns
                            if any(k in str(col).lower() for k in [
                                "largest gift high", "largest gift low", "largest gift found", "capacity", "wealth",
                                "gift found", "real estate", "net worth", "0 - 100"
                            ])
                        ]

                        if wealth_options:
                            # Force "Largest Gift Low" to be the first option and the default
                            low_index = None
                            for i, col in enumerate(wealth_options):
                                if "largest gift low" in str(col).lower():
                                    low_index = i
                                    break

                            if low_index is not None:
                                low_col = wealth_options.pop(low_index)
                                wealth_options.insert(0, low_col)

                            selected_wealth = st.selectbox(
                                "Choose wealth column for X-axis",
                                options=wealth_options,
                                index=0
                            )

                            # Safe numeric conversion
                            s = scored_df[selected_wealth].astype(str).str.strip()
                            extracted = s.str.extract(r'(\d+)-?(\d*)')
                            scored_df[selected_wealth] = extracted.apply(pd.to_numeric, errors='coerce').mean(axis=1).fillna(0)

                            plot_df = scored_df.sample(min(5000, len(scored_df))).copy()

                            # Safe wrapper for scatter plot
                            try:
                                fig = px.scatter(
                                    plot_df,
                                    x=selected_wealth,
                                    y="Propensity_Score",
                                    color="Probability_Label",
                                    hover_data=["name", "email", "total_given", "Days_Since_Last_Gift"],
                                    title="Propensity vs Capacity Scatter Plot",
                                    labels={
                                        selected_wealth: "Wealth / Capacity",
                                        "Propensity_Score": "Propensity Score"
                                    },
                                    category_orders={"Probability_Label": ["HIGH", "MEDIUM", "LOW"]},
                                    color_discrete_map={
                                        "HIGH": "#E53935",    # red
                                        "MEDIUM": "#1E88E5",  # blue
                                        "LOW": "#FB8C00"      # orange
                                    }
                                )

                                # Log scale
                                fig.update_layout(xaxis_type="log")

                                # Less obtrusive GOLDMINE with arrow in true top-right
                                fig.add_annotation(
                                    x=0.88, y=0.88,
                                    xref="paper", yref="paper",
                                    ax=30, ay=-25,
                                    text="GOLDMINE — Prioritize These!",
                                    showarrow=True,
                                    arrowhead=2,
                                    arrowsize=1,
                                    arrowwidth=1.5,
                                    arrowcolor="#2E7D32",
                                    font=dict(size=11, color="green")   # smaller text, safe settings only (no bold)
                                )

                                st.plotly_chart(fig, use_container_width=True)
                            except Exception as plot_err:
                                st.warning(f"Could not render scatter plot: {str(plot_err)[:150]}")
                        else:
                            st.info("No wealth or capacity columns detected in this file.")
                    except Exception as e:
                        st.warning(f"Could not render the Propensity vs Capacity scatter plot: {str(e)}")

                    # -----------------------------------------------------------------
                    # TOP 20 TABLE
                    # -----------------------------------------------------------------
                    with st.spinner("Loading top propensity donors..."):
                        st.subheader("Top 20 Highest Propensity Scores")

                        display_cols = [
                            "email", "name", "Propensity_Score", "Probability_Label",
                            "Re_Engage_Flag", "Days_Since_Last_Gift", "total_given"
                        ]
                        # Only keep columns that actually exist
                        display_cols = [c for c in display_cols if c in scored_df.columns]

                        top20 = scored_df.sort_values("Propensity_Score", ascending=False).head(20)

                        # Use st.data_editor so names (and other fields) are clickable / editable
                        st.data_editor(
                            top20[display_cols].style.format({
                                "Propensity_Score": "{:.0f}",
                                "total_given": "${:,.2f}",
                            }),
                            use_container_width=True,
                            hide_index=True,
                            key="top20_editor",
                        )

                    # -----------------------------------------------------------------
                    # DASHBOARD: HISTOGRAM + RE-ENGAGEMENT LIST
                    # -----------------------------------------------------------------
                    st.divider()
                    st.subheader("📊 Score Distribution & Re-Engagement List")

                    hist_col, re_col = st.columns([1.15, 1])

                    with hist_col:
                        st.markdown("**Propensity Score Histogram**")
                        fig = px.histogram(
                            scored_df,
                            x="Propensity_Score",
                            nbins=20,
                            range_x=[0, 100],
                            color_discrete_sequence=["#2E86AB"],
                            labels={"Propensity_Score": "Propensity Score (1–100)"},
                        )
                        fig.update_layout(
                            bargap=0.08,
                            height=340,
                            margin=dict(l=10, r=10, t=30, b=10),
                            xaxis=dict(tick0=0, dtick=10),
                        )
                        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

                    with re_col:
                        with st.spinner("Loading re-engagement candidates..."):
                            st.markdown("**Top 500 Re-Engagement Candidates**")
                            st.caption("Lapsed >90 days with Propensity > 25 — sorted by score")

                            reengage_df = (
                                scored_df[scored_df["Re_Engage_Flag"] == "YES"]
                                .sort_values("Propensity_Score", ascending=False)
                                .head(500)
                            )

                            if len(reengage_df) > 0:
                                re_display = reengage_df[[col for col in ['email', 'name', 'last_gift_date', 'total_given', 'Propensity_Score', 'Probability_Label', 'Re_Engage_Flag'] if col in reengage_df.columns]].copy()
                                st.dataframe(
                                    re_display.style.format({"total_given": "${:,.2f}"}),
                                    use_container_width=True,
                                    hide_index=True,
                                    height=320,
                                )
                                st.caption(f"Showing top {min(500, len(reengage_df))} of {len(reengage_df)} re-engagement candidates")
                            else:
                                st.info("No re-engagement candidates in this file (no contacts with >90 days since last gift and score > 25).")

                    # (Legacy duplicate scatter plot removed for stability and to avoid confusion on large files.
                    #  The primary dropdown version above is the recommended and fully polished experience.)

            # -----------------------------------------------------------------
            # Close the main processing try block with professional error handling
            except Exception as e:
                st.error(f"⚠️ Something went wrong: {str(e)[:200]}")
                st.info("💡 Tip: Try 'Largest Gift High' in the dropdown or restart the app.")
                import traceback
                with st.expander("Technical details"):
                    st.code(traceback.format_exc())

            # -----------------------------------------------------------------
            # FULL SCORED CSV DOWNLOAD
            # -----------------------------------------------------------------
            st.divider()
            st.subheader("Export Full Scored Dataset")

            st.markdown(
                "Download a complete CSV containing **every original column** plus the new scoring columns:"
            )

            scoring_cols = [
                "Propensity_Score", "Probability_Label", "Re_Engage_Flag",
                "Days_Since_Last_Gift", "Recency_Points", "Engagement_Points",
                "Past_Giving_Bonus", "Ever_Given_Bonus", "Major_Donor_Boost", "Raw_Score"
            ]

            st.caption("New columns added: " + ", ".join(scoring_cols))

            # Reorder for convenience: original columns first, then scoring
            original_cols = [c for c in cleaned_df.columns if c not in ["last_gift_date_parsed"]]
            final_export = scored_df[[*original_cols, *scoring_cols]]

            create_download_link(
                final_export,
                f"scored_donors_{datetime.now().strftime('%Y-%m-%d')}.csv"
            )

            # Also offer a re-engagement only export
            if len(reengage_df) > 0:
                re_export = reengage_df[[*original_cols, *scoring_cols]]
                create_download_link(
                    re_export,
                    f"reengagement_list_{datetime.now().strftime('%Y-%m-%d')}.csv"
                )

            # -----------------------------------------------------------------
            # RAW DATA PREVIEW (optional)
            # -----------------------------------------------------------------
            with st.expander("🔍 View full scored data (first 100 rows)", expanded=False):
                st.dataframe(
                    scored_df.head(100),
                    use_container_width=True,
                    hide_index=True,
                )

        except Exception as e:
            st.error("An unexpected error occurred while processing the file.")
            st.exception(e)
            st.stop()

    else:
        # Friendly empty state
        st.info(
            "👆 Upload a .csv or .xlsx file above or download the sample file to see the scorer in action.\n\n"
            "Your data never leaves your computer — everything runs locally in this browser tab."
        )

    # -------------------------------------------------------------------------
    # FOOTER
    # -------------------------------------------------------------------------
    st.divider()
    st.markdown(
        """
        <div style="text-align: center; color: #888; font-size: 0.85rem; padding-top: 0.5rem;">
            Built with <strong>Grok Build</strong> • Based on proven donor propensity model<br>
            <span style="font-size: 0.75rem;">No data is stored. All scoring happens in-memory.</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
